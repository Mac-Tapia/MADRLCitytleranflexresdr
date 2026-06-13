"""Create the required Excel workbook template for the MADRL CityLearn review."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEETS = [
    "Matriz_50_investigaciones",
    "Resumen_ejecutivo",
    "KPIs_y_metricas",
    "Marco_metodologico_MADRL",
    "CityLearn_v3_Propuesto",
    "Backends_MADRL",
    "MARLlib_Integracion",
    "CityLearn_CO2_Costos",
    "Datasets_y_codigo",
    "Lectura_priorizada",
    "Cadenas_de_busqueda",
    "Glosario_MADRL",
    "Arquitectura_Propuesta",
    "Aplicabilidad_SEAI_Iquitos",
]

MATRIX_COLUMNS = [
    "N.º", "Año", "Idioma", "Tipo de documento", "Título de la investigación",
    "Autor(es)", "Universidad, revista o congreso", "Indexación o fuente",
    "País o contexto de estudio", "Palabras clave asociadas",
    "Relación con CityLearn v2", "Relación con CityLearn v3 propuesto",
    "Relación con MADRL", "Relación con MARLlib",
    "MARLlib usado directamente: sí/no/parcial",
    "MARLlib como referencia metodológica", "Compatibilidad con MARLlib",
    "Tipo de integración posible con CityLearn v2", "Requiere wrapper personalizado",
    "Requiere adaptación a Dec-POMDP", "Requiere adaptación CTDE",
    "Requiere backend personalizado", "Problema de investigación",
    "Objetivo de investigación", "Variables de investigación",
    "Variable independiente", "Variable dependiente", "Variables de control",
    "Nivel de investigación", "Diseño de investigación", "Metodología empleada",
    "Algoritmo o modelo usado", "Backend asociado: HAPPO, MASAC, MATD3, MAAC u otro",
    "Tipo de cooperación", "CTDE: sí/no/parcial",
    "Modelo formal: MMDP, Dec-POMDP u otro",
    "Observabilidad: total, parcial o no especificada", "Estado global usado",
    "Observaciones locales usadas", "Acciones de los agentes", "Función de recompensa",
    "Recompensa individual, compartida o híbrida", "Enfoque multiobjetivo",
    "Enfoque multicriterio", "Uso de Optuna o ajuste de hiperparámetros",
    "Hiperparámetros ajustados", "Métricas de entrenamiento MADRL",
    "Métricas de convergencia", "Métricas de robustez", "Métricas de estabilidad",
    "Entorno virtual o simulador usado", "Dataset usado", "Link o ubicación del dataset",
    "Variables del dataset", "GitHub o repositorio de código", "Link del PDF o artículo",
    "DOI o enlace académico", "KPIs de flexibilidad energética",
    "KPIs de emisiones de CO₂", "KPIs de costos energéticos",
    "KPIs de respuesta de demanda", "KPIs de resiliencia eléctrica",
    "Resultados principales", "Resultados cuantitativos", "Aporte a la ciencia",
    "Conclusiones principales", "Limitaciones", "Aplicabilidad a CityLearn v3 propuesto",
    "Aplicabilidad a sistemas eléctricos aislados", "Aplicabilidad al SEAI Iquitos",
    "Relación con PV, BESS o EV charging", "Utilidad para la tesis",
    "Prioridad de lectura", "Observaciones de verificación",
]

GENERIC_COLUMNS = ["Sección", "Contenido", "Fuente/Evidencia", "Observaciones"]


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, max_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)


def build_workbook(output: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name in SHEETS:
        ws = wb.create_sheet(sheet_name)
        headers = MATRIX_COLUMNS if sheet_name == "Matriz_50_investigaciones" else GENERIC_COLUMNS
        ws.append(headers)
        style_header(ws)
        autosize(ws)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", required=True, help="Path to the .xlsx workbook to create.")
    args = parser.parse_args()
    build_workbook(Path(args.output))
    print(Path(args.output).resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

